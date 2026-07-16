from flask import Flask, render_template, request, redirect, send_file, session
import sqlite3
import os
import tempfile

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

app = Flask(__name__)
app.secret_key = "SRM2026"

# -----------------------------
# Base de données
# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "parc.db")


# -----------------------------
# Login
# -----------------------------
@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        if username == "srm_admin" and password == "SRM@2026":

            session["user"] = username

            return redirect("/accueil")

        else:

            return render_template(
                "login.html",
                erreur="Nom d'utilisateur ou mot de passe incorrect."
            )

    return render_template("login.html")


# -----------------------------
# Accueil
# -----------------------------
@app.route("/accueil")
def accueil():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM utilisateurs")
    nb_utilisateurs = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM ordinateurs")
    nb_ordinateurs = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM ordinateurs
        WHERE type_ordinateur='Portable'
    """)
    nb_portable = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM ordinateurs
        WHERE type_ordinateur='Bureau'
    """)
    nb_bureau = cursor.fetchone()[0]

    cursor.execute("""
        SELECT nom_prenom, service, matricule
        FROM utilisateurs
        ORDER BY id DESC
        LIMIT 5
    """)
    derniers_utilisateurs = cursor.fetchall()

    cursor.execute("""
        SELECT nom_ordinateur, numero_serie, type_ordinateur, matricule
        FROM ordinateurs
        ORDER BY id DESC
        LIMIT 5
    """)
    derniers_ordinateurs = cursor.fetchall()
    total_pc = nb_portable + nb_bureau

    conn.close()

    return render_template(
        "index.html",
        nb_utilisateurs=nb_utilisateurs,
        nb_ordinateurs=nb_ordinateurs,
        nb_portable=nb_portable,
        nb_bureau=nb_bureau,
        total_pc=total_pc,
        derniers_utilisateurs=derniers_utilisateurs,
        derniers_ordinateurs=derniers_ordinateurs
    )


# -----------------------------
# Gestion des utilisateurs
# -----------------------------
@app.route("/utilisateurs", methods=["GET", "POST"])
def utilisateurs():

    if "user" not in session:
        return redirect("/")

    if request.method == "POST":

        nom_prenom = request.form["nom_prenom"]
        matricule = request.form["matricule"]
        service = request.form["service"]

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO utilisateurs
        (nom_prenom, matricule, service)
        VALUES (?, ?, ?)
        """, (nom_prenom, matricule, service))

        conn.commit()
        conn.close()

        return redirect("/utilisateurs")

    recherche = request.args.get("recherche", "")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
    SELECT * FROM utilisateurs
    WHERE matricule LIKE ?
    ORDER BY CAST(matricule AS INTEGER) ASC
    """, ('%' + recherche + '%',))

    utilisateurs_liste = cursor.fetchall()

    conn.close()

    return render_template(
        "utilisateurs.html",
        utilisateurs=utilisateurs_liste
    )
# -----------------------------
# Modifier utilisateur
# -----------------------------
@app.route("/modifier_utilisateur/<int:id>", methods=["GET", "POST"])
def modifier_utilisateur(id):

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    if request.method == "POST":

        nom_prenom = request.form["nom_prenom"]
        matricule = request.form["matricule"]
        service = request.form["service"]

        cursor.execute("""
        UPDATE utilisateurs
        SET nom_prenom=?,
            matricule=?,
            service=?
        WHERE id=?
        """, (nom_prenom, matricule, service, id))

        conn.commit()
        conn.close()

        return redirect("/utilisateurs")

    cursor.execute(
        "SELECT * FROM utilisateurs WHERE id=?",
        (id,)
    )

    utilisateur = cursor.fetchone()

    conn.close()

    return render_template(
        "modifier_utilisateur.html",
        utilisateur=utilisateur
    )


# -----------------------------
# Supprimer utilisateur
# -----------------------------
@app.route("/supprimer_utilisateur/<int:id>")
def supprimer_utilisateur(id):

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM utilisateurs WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/utilisateurs")


# -----------------------------
# Gestion des ordinateurs
# -----------------------------
@app.route("/ordinateur", methods=["GET", "POST"])
def ordinateur():

    if "user" not in session:
        return redirect("/")

    if request.method == "POST":

        type_ordinateur = request.form["type_ordinateur"]
        nom_ordinateur = request.form["nom_ordinateur"]
        numero_serie = request.form["numero_serie"]
        matricule = request.form["matricule"]

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO ordinateurs
        (type_ordinateur, nom_ordinateur, numero_serie, matricule)
        VALUES (?, ?, ?, ?)
        """, (
            type_ordinateur,
            nom_ordinateur,
            numero_serie,
            matricule
        ))

        conn.commit()
        conn.close()

        return redirect("/ordinateur")

    recherche = request.args.get("recherche", "")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
    SELECT * FROM ordinateurs
    WHERE numero_serie LIKE ?
    ORDER BY CAST(matricule AS INTEGER) ASC
    """, ('%' + recherche + '%',))

    ordinateurs_liste = cursor.fetchall()

    conn.close()

    return render_template(
        "ordinateur.html",
        ordinateurs=ordinateurs_liste
    )


# -----------------------------
# Modifier ordinateur
# -----------------------------
@app.route("/modifier_ordinateur/<int:id>", methods=["GET", "POST"])
def modifier_ordinateur(id):

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    if request.method == "POST":

        type_ordinateur = request.form["type_ordinateur"]
        nom_ordinateur = request.form["nom_ordinateur"]
        numero_serie = request.form["numero_serie"]
        matricule = request.form["matricule"]

        cursor.execute("""
        UPDATE ordinateurs
        SET type_ordinateur=?,
            nom_ordinateur=?,
            numero_serie=?,
            matricule=?
        WHERE id=?
        """, (
            type_ordinateur,
            nom_ordinateur,
            numero_serie,
            matricule,
            id
        ))

        conn.commit()
        conn.close()

        return redirect("/ordinateur")

    cursor.execute(
        "SELECT * FROM ordinateurs WHERE id=?",
        (id,)
    )

    ordinateur = cursor.fetchone()

    conn.close()

    return render_template(
        "modifier_ordinateur.html",
        ordinateur=ordinateur
    )


# -----------------------------
# Supprimer ordinateur
# -----------------------------
@app.route("/supprimer_ordinateur/<int:id>")
def supprimer_ordinateur(id):

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM ordinateurs WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/ordinateur")
# -----------------------------
# Statistiques
# -----------------------------
@app.route("/statistiques")
def statistiques():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()


    # Nombre des utilisateurs
    cursor.execute("SELECT COUNT(*) FROM utilisateurs")
    nb_utilisateurs = cursor.fetchone()[0]

    # Nombre des ordinateurs
    cursor.execute("SELECT COUNT(*) FROM ordinateurs")
    nb_ordinateurs = cursor.fetchone()[0]

    # Nombre des PC Bureau
    cursor.execute("""
    SELECT COUNT(*)
    FROM ordinateurs
    WHERE type_ordinateur='Bureau'
    """)
    nb_bureau = cursor.fetchone()[0]

    # Nombre des PC Portable
    cursor.execute("""
    SELECT COUNT(*)
    FROM ordinateurs
    WHERE type_ordinateur='Portable'
    """)
    nb_portable = cursor.fetchone()[0]

    # Liste
    cursor.execute("""
    SELECT
        u.nom,
        u.prenom,
        u.matricule,
        u.service,
        o.type_ordinateur,
        o.nom_ordinateur,
        o.numero_serie
    FROM utilisateurs u
    LEFT JOIN ordinateurs o
    ON u.matricule = o.matricule
    ORDER BY CAST(u.matricule AS INTEGER)
    """, )

    liste = cursor.fetchall()

    conn.close()

    return render_template(
        "statistiques.html",
        nb_utilisateurs=nb_utilisateurs,
        nb_ordinateurs=nb_ordinateurs,
        nb_bureau=nb_bureau,
        nb_portable=nb_portable,
        liste=liste
    )


# -----------------------------
# Déconnexion
# -----------------------------
@app.route("/logout")
def logout():

    session.clear()

    return redirect("/")
# -----------------------------
# Export PDF
# -----------------------------
@app.route("/export_pdf")
def export_pdf():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        u.nom,
        u.prenom,
        u.matricule,
        u.service,
        o.type_ordinateur,
        o.nom_ordinateur,
        o.numero_serie
    FROM utilisateurs u
    LEFT JOIN ordinateurs o
    ON u.matricule = o.matricule
    ORDER BY CAST(u.matricule AS INTEGER)
    """)

    donnees = cursor.fetchall()

    conn.close()

    fichier = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".pdf"
    ).name

    pdf = SimpleDocTemplate(fichier)

    elements = []

    styles = getSampleStyleSheet()

    titre = Paragraph(
        "RAPPORT PARC INFORMATIQUE",
        styles["Title"]
    )

    elements.append(titre)
    elements.append(Spacer(1, 20))

    data = [[
        "Nom",
        "Prenom",
        "Matricule",
        "Service",
        "Type PC",
        "Nom PC",
        "Numero Serie"
    ]]

    for ligne in donnees:
        data.append(list(ligne))

    tableau = Table(data)

    tableau.setStyle(TableStyle([

        ('BACKGROUND', (0,0), (-1,0), colors.green),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke)

    ]))

    elements.append(tableau)

    pdf.build(elements)

    return send_file(
        fichier,
        as_attachment=True,
        download_name="rapport_parc_informatique.pdf"
    )
# -----------------------------
# Export Excel
# -----------------------------
@app.route("/export_excel")
def export_excel():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        u.nom,
        u.prenom,
        u.matricule,
        u.service,
        o.type_ordinateur,
        o.nom_ordinateur,
        o.numero_serie
    FROM utilisateurs u
    LEFT JOIN ordinateurs o
    ON u.matricule = o.matricule
    ORDER BY CAST(u.matricule AS INTEGER)
    """)

    donnees = cursor.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active

    ws.title = "Parc Informatique"

    headers = [
        "Nom",
        "Prenom",
        "Matricule",
        "Service",
        "Type PC",
        "Nom PC",
        "Numero Serie"
    ]

    ws.append(headers)

    for ligne in donnees:
        ws.append(ligne)

    header_fill = PatternFill(
        start_color="2E7D32",
        end_color="2E7D32",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 5

    fichier = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    ).name

    wb.save(fichier)

    return send_file(
        fichier,
        as_attachment=True,
        download_name="rapport_parc_informatique.xlsx"
    )


# -----------------------------
# tous utilisateurs
# -----------------------------
@app.route("/tous_utilisateurs")
def tous_utilisateurs():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT nom_prenom, matricule, service
        FROM utilisateurs
        ORDER BY id DESC
    """)

    utilisateurs = cursor.fetchall()

    conn.close()

    return render_template(
        "tous_utilisateurs.html",
        utilisateurs=utilisateurs
    )



# -----------------------------
# tous ordinateur
# -----------------------------
@app.route("/tous_ordinateurs")
def tous_ordinateurs():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT type_ordinateur,
               nom_ordinateur,
               numero_serie,
               matricule
        FROM ordinateurs
        ORDER BY id DESC
    """)

    ordinateurs = cursor.fetchall()

    conn.close()

    return render_template(
        "tous_ordinateurs.html",
        ordinateurs=ordinateurs
    )




# -----------------------------
# Lancer l'application
# -----------------------------
if __name__ == "__main__":
    app.run(debug=True)