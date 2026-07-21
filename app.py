from flask import Flask, render_template, request, redirect, send_file, session, flash
import sqlite3
import os
print(os.path.abspath("parc.db"))
import tempfile

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

app = Flask(__name__)
app.secret_key = "SRM2026"

# -----------------------------
# Base de données
# -----------------------------



# -----------------------------
# Login
# -----------------------------
@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        if username == "srm_admin" and password == "SRM@2026":

            session["user"] = username

            return redirect("/accueil")

        else:

            return render_template(
                "login.html",
                erreur="Nom d'utilisateur ou mot de passe incorrect."
            )

    return render_template("login.html")


# -----------------------------
# Accueil
# -----------------------------
@app.route("/accueil")
def accueil():
    print("Current directory:", os.getcwd())
    print("Database path:", os.path.abspath("parc.db"))

    if "user" not in session:
        return redirect("/")
    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM utilisateurs")
    nb_utilisateurs = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM ordinateurs")
    nb_ordinateurs = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM ordinateurs
        WHERE type_ordinateur='Portable'
    """)
    nb_portable = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM ordinateurs
        WHERE type_ordinateur='Bureau'
    """)
    nb_bureau = cursor.fetchone()[0]

    cursor.execute("""
        SELECT nom_prenom, service, matricule
        FROM utilisateurs
        ORDER BY id DESC
        LIMIT 5
    """)
    derniers_utilisateurs = cursor.fetchall()

    cursor.execute("""
        SELECT nom_ordinateur, numero_serie, type_ordinateur, matricule
        FROM ordinateurs
        ORDER BY id DESC
        LIMIT 5
    """)
    derniers_ordinateurs = cursor.fetchall()
    total_pc = nb_portable + nb_bureau

    conn.close()

    return render_template(
        "index.html",
        nb_utilisateurs=nb_utilisateurs,
        nb_ordinateurs=nb_ordinateurs,
        nb_portable=nb_portable,
        nb_bureau=nb_bureau,
        total_pc=total_pc,
        derniers_utilisateurs=derniers_utilisateurs,
        derniers_ordinateurs=derniers_ordinateurs
    )


# -----------------------------
# Gestion des utilisateurs
# -----------------------------
@app.route("/utilisateurs", methods=["GET", "POST"])
def utilisateurs():

    if "user" not in session:
        return redirect("/")

    if request.method == "POST":

        nom_prenom = request.form["nom_prenom"]
        matricule = request.form["matricule"]
        service = request.form["service"]

        conn = sqlite3.connect("parc.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM utilisateurs WHERE matricule = ?",
            (matricule,)
            )
        existe = cursor.fetchone()
        if existe:
            conn.close()
            flash("Erreur : ce matricule existe déjà. Veuillez saisir un matricule différent.", "error")
            return redirect("/utilisateurs")

        cursor.execute("""
                      INSERT INTO utilisateurs
                       (nom_prenom, matricule, service)
                       VALUES (?, ?, ?)
                       """, (nom_prenom, matricule, service))

        conn.commit()
        conn.close()
        
        flash("Utilisateur ajouté avec succès.", "success")
        return redirect("/utilisateurs")

    recherche = request.args.get("recherche", "")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    cursor.execute("""
    SELECT * FROM utilisateurs
    WHERE matricule LIKE ?
    ORDER BY CAST(matricule AS UNSIGNED) ASC
    """, ('%' + recherche + '%',))

    utilisateurs_liste = cursor.fetchall()

    conn.close()

    return render_template(
        "utilisateurs.html",
        utilisateurs=utilisateurs_liste
    )
# -----------------------------
# Modifier utilisateur
# -----------------------------
@app.route("/modifier_utilisateur/<int:id>", methods=["GET", "POST"])
def modifier_utilisateur(id):

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    if request.method == "POST":

        nom_prenom = request.form["nom_prenom"]
        matricule = request.form["matricule"]
        service = request.form["service"]

        cursor.execute("""
        UPDATE utilisateurs
        SET nom_prenom=?,
            matricule=?,
            service=?
        WHERE id=?
        """, (nom_prenom, matricule, service, id))

        conn.commit()
        conn.close()

        return redirect("/utilisateurs")

    cursor.execute(
        "SELECT * FROM utilisateurs WHERE id=?",
        (id,)
    )

    utilisateur = cursor.fetchone()

    conn.close()

    return render_template(
        "modifier_utilisateur.html",
        utilisateur=utilisateur
    )


# -----------------------------
# Supprimer utilisateur
# -----------------------------
@app.route("/supprimer_utilisateur/<int:id>")
def supprimer_utilisateur(id):

    if "user" not in session:
        return redirect("/")
    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM utilisateurs WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/utilisateurs")


# -----------------------------
# Gestion des ordinateurs
# -----------------------------
@app.route("/ordinateur", methods=["GET", "POST"])
def ordinateur():

    if "user" not in session:
        return redirect("/")

    if request.method == "POST":

        type_ordinateur = request.form["type_ordinateur"]
        nom_ordinateur = request.form["nom_ordinateur"]
        numero_serie = request.form["numero_serie"]
        matricule = request.form["matricule"]

        conn = sqlite3.connect("parc.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM utilisateurs WHERE matricule = ?",
            (matricule,)
        )
        utilisateur = cursor.fetchone()
        if not utilisateur:
            conn.close()
            flash("Erreur : aucun utilisateur ne possède ce matricule. Veuillez d'abord créer l'utilisateur.", "error")
            return redirect("/ordinateur")

        cursor.execute("""
        INSERT INTO ordinateurs
        (type_ordinateur, nom_ordinateur, numero_serie, matricule)
        VALUES (?, ?, ?, ?)
        """, (
            type_ordinateur,
            nom_ordinateur,
            numero_serie,
            matricule
        ))

        conn.commit()
        conn.close()
        
        flash("Ordinateur ajouté avec succès.", "success")
        return redirect("/ordinateur")

    recherche_numero = request.args.get("recherche_numero", "")
    recherche_matricule = request.args.get("recherche_matricule", "")
    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()
    if recherche_numero:
        cursor.execute("""
        SELECT * FROM ordinateurs
        WHERE numero_serie LIKE ?
        ORDER BY CAST(matricule AS INTEGER) ASC
    """, ('%' + recherche_numero + '%',))

    elif recherche_matricule:
        cursor.execute("""
        SELECT * FROM ordinateurs
        WHERE matricule LIKE ?
        ORDER BY CAST(matricule AS INTEGER) ASC
    """, ('%' + recherche_matricule + '%',))

    else:
        cursor.execute("""
        SELECT * FROM ordinateurs
        ORDER BY CAST(matricule AS INTEGER) ASC
    """)

    ordinateurs_liste = cursor.fetchall()

    conn.close()
    return render_template(
        "ordinateur.html",
        ordinateurs=ordinateurs_liste
    )


# -----------------------------
# Modifier ordinateur
# -----------------------------
@app.route("/modifier_ordinateur/<int:id>", methods=["GET", "POST"])
def modifier_ordinateur(id):

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    if request.method == "POST":

        type_ordinateur = request.form["type_ordinateur"]
        nom_ordinateur = request.form["nom_ordinateur"]
        numero_serie = request.form["numero_serie"]
        matricule = request.form["matricule"]

        cursor.execute("""
        UPDATE ordinateurs
        SET type_ordinateur=?,
            nom_ordinateur=?,
            numero_serie=?,
            matricule=?
        WHERE id=?
        """, (
            type_ordinateur,
            nom_ordinateur,
            numero_serie,
            matricule,
            id
        ))

        conn.commit()
        conn.close()

        return redirect("/ordinateur")

    cursor.execute(
        "SELECT * FROM ordinateurs WHERE id=?",
        (id,)
    )

    ordinateur = cursor.fetchone()

    conn.close()

    return render_template(
        "modifier_ordinateur.html",
        ordinateur=ordinateur
    )


# -----------------------------
# Supprimer ordinateur
# -----------------------------
@app.route("/supprimer_ordinateur/<int:id>")
def supprimer_ordinateur(id):

    if "user" not in session:
        return redirect("/")
    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    

    cursor.execute(
        "DELETE FROM ordinateurs WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect("/ordinateur")
# -----------------------------
# Statistiques
# -----------------------------
@app.route("/statistiques")
def statistiques():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()


    # Nombre des utilisateurs
    cursor.execute("SELECT COUNT(*) FROM utilisateurs")
    nb_utilisateurs = cursor.fetchone()[0]

    # Nombre des ordinateurs
    cursor.execute("SELECT COUNT(*) FROM ordinateurs")
    nb_ordinateurs = cursor.fetchone()[0]

    # Nombre des PC Bureau
    cursor.execute("""
    SELECT COUNT(*)
    FROM ordinateurs
    WHERE type_ordinateur='Bureau'
    """)
    nb_bureau = cursor.fetchone()[0]

    # Nombre des PC Portable
    cursor.execute("""
    SELECT COUNT(*)
    FROM ordinateurs
    WHERE type_ordinateur='Portable'
    """)
    nb_portable = cursor.fetchone()[0]

    # Liste
    cursor.execute("""
    SELECT
        u.nom_prenom,
        u.matricule,
        u.service,
        o.type_ordinateur,
        o.nom_ordinateur,
        o.numero_serie
    FROM utilisateurs u
    LEFT JOIN ordinateurs o
    ON u.matricule = o.matricule
    ORDER BY CAST(u.matricule AS INTEGER)
    """, )

    liste = cursor.fetchall()

    conn.close()

    return render_template(
        "statistiques.html",
        nb_utilisateurs=nb_utilisateurs,
        nb_ordinateurs=nb_ordinateurs,
        nb_bureau=nb_bureau,
        nb_portable=nb_portable,
        liste=liste
    )


# -----------------------------
# Déconnexion
# -----------------------------
@app.route("/logout")
def logout():

    session.clear()

    return redirect("/")
# -----------------------------
# Export PDF
# -----------------------------
@app.route("/telecharger_pdf")
def telecharger_pdf():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        u.nom_prenom,
        u.matricule,
        u.service,
        o.type_ordinateur,
        o.nom_ordinateur,
        o.numero_serie
    FROM utilisateurs u
    LEFT JOIN ordinateurs o
    ON u.matricule = o.matricule
    ORDER BY CAST(u.matricule AS INTEGER)
    """)

    donnees = cursor.fetchall()

    conn.close()

    data = [[
        "Nom et prénom",
        "Matricule",
        "Service",
        "Type PC",
        "Nom PC",
        "Numéro Série"
    ]]

    for ligne in donnees:
        data.append(list(ligne))

    fichier = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".pdf"
    ).name

    pdf = SimpleDocTemplate(fichier)

    table = Table(data)

    table.setStyle(TableStyle([

        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

        ('GRID', (0, 0), (-1, -1), 1, colors.black),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('ALIGN', (0, 0), (-1, -1), 'CENTER')

    ]))

    pdf.build([table])

    return send_file(
        fichier,
        as_attachment=True,
        download_name="rapport_parc_informatique.pdf",
        mimetype="application/pdf"
    )
# -----------------------------
# Export Excel
# -----------------------------
@app.route("/export_excel")
def export_excel():

    if "user" not in session:
        return redirect("/")

    return render_template("export_excel.html")
@app.route("/telecharger_excel")
def telecharger_excel():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        u.nom_prenom, 
        u.matricule,
        u.service,
        o.type_ordinateur,
        o.nom_ordinateur,
        o.numero_serie
    FROM utilisateurs u
    LEFT JOIN ordinateurs o
    ON u.matricule = o.matricule
    ORDER BY u.service ASC
    """) 

    donnees = cursor.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active

    ws.title = "Parc Informatique"

    headers = [
        "Nom et prenom",
        "Matricule",
        "Service",
        "Type PC",
        "Nom PC",
        "Numero Serie"
    ]

    ws.append(headers)

    for ligne in donnees:
        ws.append(ligne)
    header_fill = PatternFill(
        start_color="2E7D32",
        end_color="2E7D32",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 5

    fichier = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    ).name

    wb.save(fichier)

    return send_file(
        fichier,
        as_attachment=True,
        download_name="rapport_parc_informatique.xlsx"
    )

# -----------------------------
# tous utilisateurs
# -----------------------------
@app.route("/tous_utilisateurs")
def tous_utilisateurs():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT nom_prenom, matricule, service
        FROM utilisateurs
        ORDER BY id DESC
    """)

    utilisateurs = cursor.fetchall()

    conn.close()

    return render_template(
        "tous_utilisateurs.html",
        utilisateurs=utilisateurs
    )



# -----------------------------
# tous ordinateur
# -----------------------------
@app.route("/tous_ordinateurs")
def tous_ordinateurs():

    if "user" not in session:
        return redirect("/")

    conn = sqlite3.connect("parc.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT type_ordinateur,
               nom_ordinateur,
               numero_serie,
               matricule
        FROM ordinateurs
        ORDER BY id DESC
    """)

    ordinateurs = cursor.fetchall()

    conn.close()

    return render_template(
        "tous_ordinateurs.html",
        ordinateurs=ordinateurs
    )


# -----------------------------
# assistant AI
# -----------------------------
@app.route("/assistance", methods=["GET", "POST"])
def assistance():

    if "user" not in session:
        return redirect("/")

    if "conversation" not in session:
        session["conversation"] = []

    conversation = session["conversation"]

    if request.method == "POST":

        question = request.form["question"].lower()
        conn = sqlite3.connect("parc.db")
        cursor = conn.cursor()

       


# ==========================
# NOMBRE D'UTILISATEURS
# ==========================

        if (
            "combien" in question
            and "utilisateur" in question
            and "service" not in question
            ):
            cursor.execute("SELECT COUNT(*) FROM utilisateurs")
            nb = cursor.fetchone()[0]
            reponse = f"Il y a {nb} utilisateurs."


# ==========================
# NOMBRE D'ORDINATEURS
# ==========================
        elif "combien" in question and "ordinateur" in question:
            cursor.execute("SELECT COUNT(*) FROM ordinateurs")
            nb = cursor.fetchone()[0]
            reponse = f"Il y a {nb} ordinateurs."


# ==========================
# LISTE DES UTILISATEURS
# ==========================

        elif "liste" in question and "utilisateur" in question:
            cursor.execute("""
            SELECT nom_prenom
        FROM utilisateurs
        ORDER BY nom_prenom
    """)
            resultat = cursor.fetchall()
            if resultat:
                 reponse = "Liste des utilisateurs :<br><br>"
                 for r in resultat:
                    reponse += f"• {r[0]}<br>"
            else:
                reponse = "Aucun utilisateur trouvé."


# ==========================
# LISTE DES ORDINATEURS
# ==========================

        elif "liste" in question and "ordinateur" in question:
            cursor.execute("""
        SELECT nom_ordinateur
        FROM ordinateurs
        ORDER BY nom_ordinateur
    """)
            resultat = cursor.fetchall()

            if resultat:

                reponse = "Liste des ordinateurs :<br><br>"
                for r in resultat:
                    reponse += f"• {r[0]}<br>"
                else:
                     reponse = "Aucun ordinateur trouvé."
# ==========================
# MATRICULE D'UN UTILISATEUR
# ==========================

        elif "matricule de" in question:
            cursor.execute("""
        SELECT nom_prenom, matricule
        FROM utilisateurs
    """)
            utilisateurs = cursor.fetchall()
            trouve = False
            for u in utilisateurs:
                if u[0] and u[0].lower() in question:
                    reponse = f"La matricule de {u[0]} est {u[1]}."
                    trouve = True
                    break
            if not trouve:
                    reponse = "Utilisateur introuvable."


# ==========================
# À QUI APPARTIENT LA MATRICULE
# ==========================

        elif "matricule" in question:
            cursor.execute("""
        SELECT nom_prenom, matricule
        FROM utilisateurs
    """)
            utilisateurs = cursor.fetchall()
            trouve = False
            for u in utilisateurs:
                if str(u[1]) in question:
                    reponse = f"La matricule {u[1]} appartient à {u[0]}."
                    trouve = True
                    break
            if not trouve:
                    reponse = "Matricule introuvable."


# ==========================
# INFORMATIONS SUR UN UTILISATEUR
# ==========================

        elif "informations sur" in question:
            cursor.execute("""
        SELECT nom_prenom, matricule, service
        FROM utilisateurs
    """)
            utilisateurs = cursor.fetchall()
            trouve = False
            for u in utilisateurs:
                 if u[0] and u[0].lower() in question:
                    reponse = (
                    f"Nom : {u[0]}<br>"
                f"Matricule : {u[1]}<br>"
                f"Service : {u[2]}"
                )
                    trouve = True
                    break
            if not trouve:
                     reponse = "Utilisateur introuvable."
# ==========================
# COMBIEN DE SERVICES
# ==========================

        elif "combien de services" in question or "nombre de services" in question:

            cursor.execute("""
        SELECT COUNT(DISTINCT service)
        FROM utilisateurs
    """)
            nb = cursor.fetchone()[0]
            reponse = f"Il existe {nb} service(s)."


# ==========================
# SERVICE AVEC LE PLUS D'UTILISATEURS
# ==========================

        elif "plus d'utilisateurs" in question:
            cursor.execute("""
        SELECT service, COUNT(*) AS nb
        FROM utilisateurs
        GROUP BY service
        ORDER BY nb DESC
        LIMIT 1
    """)
            resultat = cursor.fetchone()
            if resultat:
                reponse = f"Le service {resultat[0]} contient le plus d'utilisateurs avec {resultat[1]} utilisateur(s)."
            else:
                reponse = "Aucune donnée disponible."


# ==========================
# COMBIEN D'UTILISATEURS DANS UN SERVICE
# ==========================

        elif (
            ("combien" in question or "effectif" in question)
            and "service" in question
            ):
            cursor.execute("""
                           SELECT DISTINCT service
                           FROM utilisateurs
                           """)
            services = cursor.fetchall()
            service_trouve = None
            for s in services:
                if s[0] and s[0].lower() in question:
                    service_trouve = s[0]
                    break
            if service_trouve:
                cursor.execute("""
                                   SELECT COUNT(*)
                                   FROM utilisateurs
                                    WHERE LOWER(service)=?
                                   """, (service_trouve.lower(),))
                nb = cursor.fetchone()[0]
                reponse = f"Le service {service_trouve} contient {nb} utilisateur(s)."
            else:
                    reponse = "Service introuvable."


# ==========================
# SERVICE D'UN UTILISATEUR
# ==========================

        elif "service" in question and "utilisateur" not in question:

            cursor.execute("""
        SELECT nom_prenom, service
        FROM utilisateurs
    """)
            utilisateurs = cursor.fetchall()
            trouve = False
            for u in utilisateurs:
                if u[0] and u[0].lower() in question:
                    reponse = f"{u[0]} travaille dans le service {u[1]}."
                    trouve = True
                    break
            if not trouve:
                    reponse = "Utilisateur introuvable."


# ==========================
# UTILISATEURS D'UN SERVICE
# ==========================

        elif "travaille" in question:

            cursor.execute("""
        SELECT DISTINCT service
        FROM utilisateurs
    """)
            services = cursor.fetchall()
            service_trouve = None
            for s in services:
                if s[0] and s[0].lower() in question:
                    service_trouve = s[0]
                    break
            if service_trouve:
                    cursor.execute("""
            SELECT nom_prenom
            FROM utilisateurs
            WHERE LOWER(service)=?
                                       """, (service_trouve.lower(),))
                    resultat = cursor.fetchall()
                    if resultat:
                        reponse = f"Utilisateurs du service {service_trouve} :<br><br>"
                        for r in resultat:
                            reponse += f"• {r[0]}<br>"
                        else:
                            reponse = f"Aucun utilisateur dans le service {service_trouve}."
                    else:
                        reponse = "Service introuvable."
# ==========================
# QUEL ORDINATEUR UTILISE UN UTILISATEUR
# ==========================

        elif "ordinateur" in question and "utilise" in question:

            cursor.execute("""
        SELECT nom_prenom, matricule
        FROM utilisateurs
    """)
            utilisateurs = cursor.fetchall()
            trouve = False
            for u in utilisateurs:
                if u[0] and u[0].lower() in question:
                    trouve = True
                    cursor.execute("""
                SELECT nom_ordinateur
                FROM ordinateurs
                WHERE matricule=?
            """, (u[1],))
                    ordinateur = cursor.fetchone()
                    if ordinateur:
                        reponse = f"{u[0]} utilise l'ordinateur {ordinateur[0]}."
                    else:
                        reponse = f"{u[0]} n'a pas d'ordinateur attribué."
                        break
            if not trouve:
                        reponse = "Utilisateur introuvable."


# ==========================
# QUI UTILISE UN ORDINATEUR
# ==========================

        elif "qui utilise" in question and "ordinateur" in question:

            cursor.execute("""
        SELECT nom_ordinateur, matricule
        FROM ordinateurs
    """)
            ordinateurs = cursor.fetchall()
            trouve = False
            for o in ordinateurs:
                if o[0] and o[0].lower() in question:
                    trouve = True
                    cursor.execute("""
                SELECT nom_prenom
                FROM utilisateurs
                WHERE matricule=?
            """, (o[1],))
                    utilisateur = cursor.fetchone()
                    if utilisateur:
                        reponse = f"L'ordinateur {o[0]} est utilisé par {utilisateur[0]}."
                    else:
                        reponse = "Aucun utilisateur associé."
                        break
                    if not trouve:
                        reponse = "Ordinateur introuvable."


# ==========================
# TYPE D'ORDINATEUR
# ==========================

        elif "type" in question and "ordinateur" in question:
           cursor.execute("""
        SELECT nom_ordinateur, type_ordinateur
        FROM ordinateurs
    """)
           ordinateurs = cursor.fetchall()
           trouve = False
           for o in ordinateurs:
                if o[0] and o[0].lower() in question:
                    reponse = f"L'ordinateur {o[0]} est de type {o[1]}."
                    trouve = True
                    break
                if not trouve:
                    reponse = "Ordinateur introuvable."


# ==========================
# NUMÉRO DE SÉRIE
# ==========================

        elif "numéro de série" in question or "numero de serie" in question:

            mots = question.split()
            matricule = mots[-1]

            cursor.execute("""
                SELECT numero_serie
                FROM ordinateurs
                WHERE matricule = ?
            """, (matricule,))

            resultat = cursor.fetchone()

            if resultat:
                reponse = f"Le numéro de série du matricule {matricule} est : {resultat[0]}"
            else:
                reponse = "Aucun ordinateur trouvé pour ce matricule."


# ==========================
# QUESTION NON RECONNUE
# ==========================
        else:
            reponse = "Je n'ai pas compris votre question."
            conversation.append({
                "role": "user",
                "text": question
        })

        conversation.append({
                "role": "ai",
                "text": reponse
        })

        session["conversation"] = conversation

        conn.close()
    return render_template(
            "assistant.html",
            conversation=conversation
        )


@app.route("/test")
def test():
    return "TEST OK"


@app.route("/clear")
def clear():

    session.pop("conversation", None)

    return redirect("/assistance")

# -----------------------------
# Lancer l'application
# -----------------------------
if __name__ == "__main__":
    app.run(debug=True)